import os
import tkinter as tk
from tkinter import filedialog, messagebox
import xlrd
from openpyxl import load_workbook
import polib

def convert_to_po():
    # Open a file dialog to select an XLS or XLSX file
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xls;*.xlsx"), ("All Files", "*.*")])

    if file_path:
        # Determine the file type based on the file extension
        ext = os.path.splitext(file_path)[1]
        if ext == ".xls":
            # Open the XLS file using xlrd
            book = xlrd.open_workbook(file_path)
            sheet = book.sheet_by_index(0)

        elif ext == ".xlsx":
            # Open the XLSX file using openpyxl
            book = load_workbook(filename=file_path, read_only=True)
            sheet = book.active

        # Create a new PO file
        po = polib.POFile()

        # Iterate over the rows and extract the translations
        for row_idx in range(1, sheet.max_row + 1):
            msgid = sheet.cell(row=row_idx, column=1).value
            msgstr = sheet.cell(row=row_idx, column=2).value

            # Create a new PO entry and add it to the PO file
            entry = polib.POEntry(msgid=msgid, msgstr=msgstr)
            po.append(entry)

        # Open a file dialog to select a destination folder and filename for the PO file
        save_path = filedialog.asksaveasfilename(defaultextension=".po", filetypes=[("PO Files", "*.po"), ("All Files", "*.*")])
        if save_path:
            # Save the PO file
            po.save(save_path)
            messagebox.showinfo("Conversion complete", "The file has been converted to PO format and saved to:\n" + save_path)

if __name__ == '__main__':
    convert_to_po()
